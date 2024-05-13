from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
import threading
import random
import re
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import keyboard
import string
import requests
import time
import os
import threading
import pyautogui
import pyperclip
import platform


def Scraping(Link):
    soup = BeautifulSoup(requests.get(Link).content, "html.parser")
    productsCount = 1
    title = soup.find("h1", class_="category-header-info").text
    A.SpiritSelectionNames.append(title)
    clean_title = re.sub(r"[/]", "", title)
    sheet = A.workbook.create_sheet(title=clean_title)
    while True:
        slecetions_box = soup.find("div", class_="row row-cols-1 row-cols-md-2 row-cols-lg-3 row-cols-xxl-4 g-0")
        Links = slecetions_box.find_all("a")
        for links in Links:
            PLink = link + links["href"]
            soup1 = BeautifulSoup(requests.get(PLink).content, "html.parser")
            Name = soup1.find("h1", class_="product-title").text
            Sale = soup1.find("div", class_="badge-sale")
            
            sheet.cell(row=productsCount, column=1).value = Name
            if Sale is not None:
                sheet.cell(row=productsCount, column=3).value = "Akcija"
                Price = soup1.find("div", class_="actual")
                
                if Price is not None:
                    sheet.cell(row=productsCount, column=2).value = Price.text
                else:
                    sheet.cell(row=productsCount, column=2).value = "Not found"
            else:
                sheet.cell(row=productsCount, column=3).value = "Nav"
                Price = soup1.find("div", class_="product-price mr-1 text-right").text
                sheet.cell(row=productsCount, column=2).value = Price

            sheet.cell(row=productsCount, column=4).value = PLink
            productsCount += 1 
            print(title + " " + str(productsCount))
            
        NextPage = soup.find("a", class_="btn-next")
        if NextPage is not None and 'href' in NextPage.attrs:
            soup = BeautifulSoup(requests.get(link + NextPage["href"]).content, "html.parser")
            print("page next " + title)
        else:
            break
class A:
    workbook = Workbook()
    SpiritSelectionNames = []

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--disable-features=VizDisplayCompositor")
chrome_options.add_argument("--disable-gpu")

SpritSelectionLinks = []

ProductNames = []
ProductLinks = []
ProductPrices = []
service = Service(ChromeDriverManager().install())
link = "https://www.spiritsandwine.lv"
response = requests.get(link)
x = 0   
# Get all types of drink links
soup = BeautifulSoup(response.content, "html.parser")
a_tag = soup.find_all("a", class_="dropdown-item")
for a in a_tag:
    if a["href"][:4] != "http" and x < 39: 
        SpritSelectionLinks.append(link + a["href"])
        print(link + a["href"])
        x += 1

threads = []
# Get all drink names, prices and values of all the drinks
for section in SpritSelectionLinks:
    Functions = threading.Thread(target=Scraping, args=(section,))
    Functions.start()
    threads.append(Functions)

for thread in threads:
    thread.join()

A.workbook.save("SpiritAndVine.xlsx")


print("Done")


"""
# Create a new workbook
workbook = Workbook()

sheet = workbook.active
sheet.cell(row=1, column=1).value = "Hello, World!"

max_length = 0
for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1):
    for cell in row:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

column = sheet.column_dimensions[f'{chr(0 + 65)}']
column.width = max_length

workbook.save('output.xlsx')
os.system(f'start excel {"output.xlsx"}')"""
