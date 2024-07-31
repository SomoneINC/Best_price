from openpyxl import Workbook
from selenium import webdriver
import threading
import re
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
import requests
import threading

class A:
    workbook = Workbook()
def Scraping(Link):
    soup = BeautifulSoup(requests.get(Link).content, "html.parser")
    SectionTitle = soup.find("h1", class_="page-title").text
    clean_title = SectionTitle.replace(" ", "").replace("\n", "")
    ProductCount = 1

    # Check if a sheet with the same name already exists
    sheet_title = clean_title
    sheet_count = 1
    while sheet_title in A.workbook.sheetnames:
        sheet_title = f"{clean_title} ({sheet_count})"
        sheet_count += 1

    sheet = A.workbook.create_sheet(title=sheet_title)

    while True:
        ProductList = []
        AllItems = soup.find_all("li", class_="item product product-item")
        for item in AllItems:
            ItemLink = item.find("a", class_="product photo product-item-photo")["href"]
            ProductList.append(ItemLink)

        for product in ProductList:
            soup2 = BeautifulSoup(requests.get(product, timeout=10).content, "html.parser")

            ProductTitle = soup2.find("h1", class_="page-title").text
            sheet.cell(row=ProductCount, column=1).value = ProductTitle
            productInfoMain = soup2.find("div", class_="product-info-main")
            ProductOnSalePrice = productInfoMain.find("span", class_="special-price")
            if ProductOnSalePrice is not None:
                ProductPrice = ProductOnSalePrice.find("span", class_="price").text
                
                sheet.cell(row=ProductCount, column=3).value = "Akcija"
            else:
                ProductPrice = soup2.find("span", class_="price").text
                sheet.cell(row=ProductCount, column=3).value = "Nav"
            sheet.cell(row=ProductCount, column=2).value = ProductPrice
            sheet.cell(row=ProductCount, column=4).value = product
            print(f"Scraped product {ProductCount}: {ProductTitle}")
            ProductCount += 1

        PageSelector = soup.find("div", class_="pages")
        if PageSelector is not None:
            NextButton = PageSelector.find("a", class_="action next")
            if NextButton is not None:
                soup = BeautifulSoup(requests.get(NextButton["href"]).content, "html.parser")
            else:
                break
        else:
            break

    # Add a delay to avoid exceeding the API rate limit
    time.sleep(1)



chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--disable-infobars")
chrome_options.add_argument("--disable-features=VizDisplayCompositor")
chrome_options.add_argument("--disable-gpu")

service = Service(ChromeDriverManager().install())
link = "https://alkoutlet.lv/"
response = requests.get(link)
SectionLinksList = []

soup = BeautifulSoup(response.content, "html.parser")
SelectionA = soup.find_all("div", class_="menu-item top-level category has-children")
for a in SelectionA[:-1]:
    SectionLinksBox = a.find("div", class_="menu-item-content")
    SectionLinks = SectionLinksBox.find_all("a", class_="title level2")
    for b in SectionLinks:
        if b["href"].count("/") == 4:
            SectionLinksList.append(b["href"])

threads = []
unique_strings = set()
for links in SectionLinksList:
    if links not in unique_strings:
        unique_strings.add(links)
        SectionLinksList.remove(links)


for multi in SectionLinksList:
    Function = threading.Thread(target=Scraping, args=(multi,))
    Function.start()
    threads.append(Function)

for thread in threads:
    thread.join()

A.workbook.save("AlkhOutlet.xlsx")